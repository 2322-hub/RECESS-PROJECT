import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelReportGenerator:
    """Generate Excel reports from dashboard data."""

    @staticmethod
    def generate(dashboard_data: dict, title: str = "BI Dashboard Report") -> bytes:
        wb = Workbook()

        ws = wb.active
        ws.title = "KPIs"
        kpis = dashboard_data.get("kpis", {})
        ws.append(["BI Dashboard Report"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])
        ws.append(["Metric", "Value"])
        for row in ws.iter_rows(min_row=4, max_row=4, max_col=2):
            for cell in row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER

        kpi_rows = [
            ("Total Revenue", f"UGX {kpis.get('total_revenue', 0):,.2f}"),
            ("Total Profit", f"UGX {kpis.get('total_profit', 0):,.2f}"),
            ("Profit Margin", f"{kpis.get('profit_margin', 0):.1f}%"),
            ("Total Cost", f"UGX {kpis.get('total_cost', 0):,.2f}"),
            ("Records", f"{kpis.get('record_count', 0):,}"),
        ]
        for label, value in kpi_rows:
            ws.append([label, value])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25

        ws2 = wb.create_sheet("Revenue by Region")
        rb = dashboard_data.get("revenue_breakdown", {})
        ws2.append(["Region", "Revenue (UGX)"])
        for cell in ws2[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for item in rb.get("region", []):
            ws2.append([item["label"], item["value"]])
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 20

        ws3 = wb.create_sheet("Revenue by Category")
        ws3.append(["Category", "Revenue (UGX)"])
        for cell in ws3[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for item in rb.get("product_category", []):
            ws3.append([item["label"], item["value"]])
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 20

        ws4 = wb.create_sheet("Product Performance")
        ws4.append(["Category", "Product", "Revenue", "Profit", "Quantity"])
        for cell in ws4[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for p in dashboard_data.get("product_performance", []):
            ws4.append(
                [
                    p.get("product_category", ""),
                    p.get("product_name", ""),
                    p.get("total_revenue", 0),
                    p.get("profit", 0),
                    p.get("quantity", 0),
                ]
            )
        for col_letter in ["A", "B", "C", "D", "E"]:
            ws4.column_dimensions[col_letter].width = 18

        ws5 = wb.create_sheet("Monthly Trends")
        ws5.append(["Date", "Revenue", "Profit", "Quantity"])
        for cell in ws5[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for m in dashboard_data.get("monthly_trends", []):
            ws5.append([m.get("date", ""), m.get("revenue", 0), m.get("profit", 0), m.get("quantity", 0)])
        for col_letter in ["A", "B", "C", "D"]:
            ws5.column_dimensions[col_letter].width = 18

        ws6 = wb.create_sheet("Regional Comparison")
        ws6.append(["Region", "Revenue", "Profit", "Orders", "Avg Order Value"])
        for cell in ws6[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        for r in dashboard_data.get("regional_comparison", []):
            ws6.append(
                [
                    r.get("region", ""),
                    r.get("revenue", 0),
                    r.get("profit", 0),
                    r.get("orders", 0),
                    r.get("avg_order_value", 0),
                ]
            )
        for col_letter in ["A", "B", "C", "D", "E"]:
            ws6.column_dimensions[col_letter].width = 18

        wa = dashboard_data.get("website_analytics", {})
        if wa and wa.get("total_page_views"):
            ws7 = wb.create_sheet("Website Analytics")
            ws7.append(["Metric", "Value"])
            for cell in ws7[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
            wa_rows = [
                ("Page Views", wa.get("total_page_views", 0)),
                ("Unique Visitors", wa.get("total_unique_visitors", 0)),
                ("Bounce Rate", f"{wa.get('avg_bounce_rate', 0):.1f}%"),
                ("Conversions", wa.get("total_conversions", 0)),
                ("Conversion Rate", f"{wa.get('conversion_rate', 0):.1f}%"),
            ]
            for label, value in wa_rows:
                ws7.append([label, value])
            ws7.column_dimensions["A"].width = 25
            ws7.column_dimensions["B"].width = 25

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
